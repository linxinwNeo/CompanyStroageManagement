import openpyxl
from openpyxl.styles import PatternFill

def read_xlsx_file(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Get the first sheet
    sheet = workbook.active

    # Iterate through all rows in the sheet
    for row in sheet.iter_rows(values_only=True):
        # Print each cell value in the row
        for cell in row:
            print(cell, end='\t')
        print()  # Move to the next line after printing a row


def write_to_xlsx_file(file_path):
    # Create a new workbook
    workbook = openpyxl.Workbook()

    # Get the active sheet
    sheet = workbook.active

    # Write data to cells
    sheet['A1'] = 'MODELO'
    sheet['B1'] = 'PZX X CJA'
    sheet['C1'] = 'INICIAL'
    sheet['D1'] = 'EXISTENCIAS'
    sheet['E1'] = 'VENTAS'

    # color the first row with purple
    fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
    for cell in sheet[1]:
        cell.fill = fill

    data = [
        # ('PB-0019', 60, 125, 125, 0),
        ('TG-3599-105', 360, 20, 6, 14),
    ]

    for row in data:
        sheet.append(row)

    # Save the workbook
    workbook.save(file_path)


# Example usage
# xlsx_file_path = 'output.xlsx'
# write_to_xlsx_file(xlsx_file_path)
# print(f"Excel file '{xlsx_file_path}' created successfully.")

# Example usage
xlsx_file_path = 'output.xlsx'
read_xlsx_file(xlsx_file_path)