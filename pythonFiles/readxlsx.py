import openpyxl

def read_xlsx_file(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path, data_only=True)

    # Get the first sheet
    for sheet in workbook.worksheets:
        if not sheet.iter_rows or sheet.sheet_state == 'hidden':
            continue
        # Iterate through all rows in the sheet
        for row in sheet.iter_rows():
            row_index = row[0].row
            print(row_index, row[0].value, row[1].value)
            # row_dimension = sheet.row_dimensions[row_index]
            # if row_dimension.hidden == True:
            #     continue

            # for cell in row:
            #     if cell.col_idx >= 7: break
            #     if cell.value == None: continue
            #     print(cell.value, end=" ")
            # print()
            # Print each cell value in the row
            # for cell in row:
            #     if cell != None:
            #         print(cell, end='\t')
            # print()  # Move to the next line after printing a row
        exit(1)

xlsx_file_path = 'database.xlsx'
read_xlsx_file(xlsx_file_path)